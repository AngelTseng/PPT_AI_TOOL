from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def _is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    return False


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value)


def _find_used_bounds(ws) -> tuple[int, int, int, int] | None:
    min_row = None
    min_col = None
    max_row = None
    max_col = None

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if _is_empty(cell.value):
                continue

            r, c = cell.row, cell.column
            min_row = r if min_row is None else min(min_row, r)
            min_col = c if min_col is None else min(min_col, c)
            max_row = r if max_row is None else max(max_row, r)
            max_col = c if max_col is None else max(max_col, c)

    if min_row is None:
        return None

    return min_row, max_row, min_col, max_col


def _build_nonempty_grid(ws, min_row: int, max_row: int, min_col: int, max_col: int):
    grid = []
    for r in range(min_row, max_row + 1):
        row_flags = []
        for c in range(min_col, max_col + 1):
            row_flags.append(not _is_empty(ws.cell(row=r, column=c).value))
        grid.append(row_flags)
    return grid


def _find_groups(has_content_flags: list[bool], separator_flags: list[bool]) -> list[tuple[int, int]]:
    groups: list[tuple[int, int]] = []
    start = None

    for idx, (has_content, is_sep) in enumerate(zip(has_content_flags, separator_flags)):
        if has_content and not is_sep:
            if start is None:
                start = idx
        else:
            if start is not None:
                groups.append((start, idx - 1))
                start = None

    if start is not None:
        groups.append((start, len(has_content_flags) - 1))

    return groups


def _split_blocks(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> list[tuple[int, int, int, int]]:
    grid = _build_nonempty_grid(ws, min_row, max_row, min_col, max_col)
    row_count = len(grid)
    col_count = len(grid[0]) if grid else 0

    row_nonempty_counts = [sum(1 for flag in grid[r] if flag) for r in range(row_count)]
    col_nonempty_counts = [sum(1 for r in range(row_count) if grid[r][c]) for c in range(col_count)]

    row_has_content = [count > 0 for count in row_nonempty_counts]
    col_has_content = [count > 0 for count in col_nonempty_counts]

    row_separator = [count <= max(0, int(col_count * 0.05)) for count in row_nonempty_counts]
    col_separator = [count <= max(0, int(row_count * 0.05)) for count in col_nonempty_counts]

    row_groups = _find_groups(row_has_content, row_separator)
    col_groups = _find_groups(col_has_content, col_separator)

    if not row_groups or not col_groups:
        return [(min_row, max_row, min_col, max_col)]

    blocks = []
    for r0, r1 in row_groups:
        for c0, c1 in col_groups:
            has_cell = False
            for rr in range(r0, r1 + 1):
                for cc in range(c0, c1 + 1):
                    if grid[rr][cc]:
                        has_cell = True
                        break
                if has_cell:
                    break

            if has_cell:
                blocks.append((min_row + r0, min_row + r1, min_col + c0, min_col + c1))

    if not blocks:
        return [(min_row, max_row, min_col, max_col)]

    return blocks


def _first_data_row_type_score(row: list[str]) -> float:
    if not row:
        return 0.0

    score = 0.0
    for value in row:
        v = value.strip()
        if not v:
            continue
        try:
            float(v.replace(",", ""))
            score += 1.0
        except Exception:
            score += 0.0
    return score / max(1, len(row))


def _detect_header(raw_matrix: list[list[str]]) -> bool:
    if len(raw_matrix) < 2:
        return False

    first = raw_matrix[0]
    second = raw_matrix[1]

    nonempty_first = [x for x in first if x.strip()]
    if not nonempty_first:
        return False

    first_string_ratio = sum(1 for x in nonempty_first if not x.replace(",", "").replace(".", "", 1).isdigit()) / len(nonempty_first)
    second_data_ratio = _first_data_row_type_score(second)

    return first_string_ratio >= 0.6 and second_data_ratio >= 0.2


def _extract_block_matrix(ws, r0: int, r1: int, c0: int, c1: int) -> list[list[str]]:
    matrix = []
    for r in range(r0, r1 + 1):
        row = []
        for c in range(c0, c1 + 1):
            row.append(_cell_to_text(ws.cell(row=r, column=c).value))
        matrix.append(row)

    while matrix and all(not str(x).strip() for x in matrix[-1]):
        matrix.pop()

    if not matrix:
        return []

    col_count = len(matrix[0])
    keep = [False] * col_count
    for row in matrix:
        for idx, cell in enumerate(row):
            if str(cell).strip():
                keep[idx] = True

    trimmed = []
    for row in matrix:
        trimmed.append([row[idx] for idx in range(col_count) if keep[idx]])

    return trimmed


def _to_range_str(r0: int, r1: int, c0: int, c1: int) -> str:
    return f"{get_column_letter(c0)}{r0}:{get_column_letter(c1)}{r1}"


def extract_excel_to_payload(path: str) -> dict:
    workbook_path = Path(path)
    wb = load_workbook(filename=path, data_only=True)

    payload = {
        "workbook_name": workbook_path.name,
        "sheets": [],
    }

    for ws in wb.worksheets:
        bounds = _find_used_bounds(ws)
        if bounds is None:
            continue

        min_row, max_row, min_col, max_col = bounds
        block_ranges = _split_blocks(ws, min_row, max_row, min_col, max_col)

        blocks = []
        for idx, (r0, r1, c0, c1) in enumerate(block_ranges, start=1):
            raw_matrix = _extract_block_matrix(ws, r0, r1, c0, c1)
            if not raw_matrix:
                continue

            has_header = _detect_header(raw_matrix)
            if has_header:
                columns = [x if x else f"Column {i+1}" for i, x in enumerate(raw_matrix[0])]
                rows = raw_matrix[1:] if len(raw_matrix) > 1 else []
            else:
                width = len(raw_matrix[0]) if raw_matrix else 0
                columns = [f"Column {i+1}" for i in range(width)]
                rows = raw_matrix

            blocks.append(
                {
                    "block_id": f"{ws.title}_B{idx}",
                    "range": _to_range_str(r0, r1, c0, c1),
                    "title": "",
                    "columns": columns,
                    "rows": rows,
                    "raw_matrix": raw_matrix,
                    "header_detected": has_header,
                }
            )

        if blocks:
            payload["sheets"].append({"sheet_name": ws.title, "blocks": blocks})

    return payload
